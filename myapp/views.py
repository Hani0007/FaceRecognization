from django.shortcuts import render, redirect
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login as auth_login
from django.contrib import messages
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
import cv2
import os
import numpy as np
from os import listdir
from os.path import isfile, join
from datetime import datetime, date
import openpyxl

# Path to the folder containing user images
data_path = 'C:/Users/E-TIME/Desktop/facerecog/imag'

# Path to the Excel file for attendance recording
excel_file = 'attendance.xlsx'

def index(request):
    return render(request, 'index.html')

def about(request):
    return render(request, 'about.html')

def services(request):
    return render(request, 'services.html')

def project(request):
    return render(request, 'projects.html')

def contact(request):
    return render(request, 'contact.html')

def login(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        if not username or not password:
            messages.error(request, "Username and password are required.")
            return redirect('login')
        
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            auth_login(request, user)
            return redirect('index')
        else:
            messages.error(request, "Invalid username or password. Please try again.")
            return redirect('login')

    return render(request, 'login.html')

def register(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')

        if User.objects.filter(username=username).exists():
            error_message = "Username already exists. Please choose a different username."
            return render(request, 'register.html', {'error_message': error_message})

        try:
            validate_email(email)
        except ValidationError:
            error_message = "Invalid email format."
            return render(request, 'register.html', {'error_message': error_message})

        if User.objects.filter(email=email).exists():
            error_message = "Email address is already registered."
            return render(request, 'register.html', {'error_message': error_message})

        if len(password) < 8:
            error_message = "Password must be at least 8 characters long."
            return render(request, 'register.html', {'error_message': error_message})

        if password != confirm_password:
            error_message = "Passwords do not match."
            return render(request, 'register.html', {'error_message': error_message})

        user = User.objects.create_user(username=username, email=email, password=password)
        return redirect('login')

    return render(request, 'register.html')

def camera(request):
    face_classifier = cv2.CascadeClassifier('C:/Users/E-TIME/Desktop/facerecog/haarcascade_frontalface_default.xml')

    def face_detector(img, size=0.5):
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        faces = face_classifier.detectMultiScale(gray, scaleFactor=1.5, minNeighbors=8)

        if len(faces) == 0:
            return img, []

        for (x, y, w, h) in faces:
            cv2.rectangle(img, (x, y), (x+w, y+h), (0, 255, 0), 2)
            roi = img[y:y+h, x:x+w]
            roi = cv2.resize(roi, (200, 200))

        return img, roi

    def create_user_folder(user_folder):
        if not os.path.exists(user_folder):
            os.makedirs(user_folder)
            print(f"Folder '{user_folder}' created successfully.")
        else:
            print(f"Folder '{user_folder}' already exists.")

    cap = cv2.VideoCapture(0)
    cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
    cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
    cap.set(cv2.CAP_PROP_BRIGHTNESS, 0.5)

    for _ in range(30):
        _ = cap.read()

    while True:
        user_name = input("Enter the name of the user: ")

        user_folder = f'{data_path}/{user_name}/'
        create_user_folder(user_folder)

        count = 0
        while True:
            ret, frame = cap.read()
            _, face = face_detector(frame)
            if len(face) > 0:  # Check if 'face' is not an empty list
                count += 1
                face = cv2.cvtColor(face, cv2.COLOR_BGR2GRAY)
                file_name_path = os.path.join(user_folder, str(count) + '.jpg')
                cv2.imwrite(file_name_path, face)
                cv2.putText(face, str(count), (50, 50), cv2.FONT_HERSHEY_COMPLEX, 1, (0, 255, 0), 2)
                cv2.imshow('Face Cropper', face)
            else:
                print("Face not found")
                pass

            if cv2.waitKey(1) == 13 or count == 500:
                break

        choice = input("Do you want to add more images for another user? (yes/no): ")
        if choice.lower() != 'yes':
            break

    cap.release()
    cv2.destroyAllWindows()
    print('Samples Collection Completed ')
    
    return render(request, 'camera.html')

def face(request):
    def mark_attendance(object_name):
        # Get current date and time
        now = datetime.now()
        current_date = now.strftime("%Y-%m-%d")
        current_time = now.strftime("%H:%M:%S")

        # Check if the Excel file exists
        if not os.path.exists(excel_file):
            # Create a new workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Attendance"
            ws.append(["Name", "Date", "Time"])  # Include a column for time
            # Save the workbook
            wb.save(excel_file)

        # Load the workbook
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active

        # Check if attendance has already been marked for today and for the specific person
        attendance_marked = False
        for row in ws.iter_rows(values_only=True):
            if row[0] == object_name and row[1] == current_date:
                print(f"Attendance for {object_name} has already been marked today.")
                attendance_marked = True
                break

        # If attendance hasn't been marked yet for today, add the entry
        if not attendance_marked:
            # Append attendance record to the Excel sheet
            max_row = ws.max_row
            ws.cell(row=max_row + 1, column=1, value=object_name)
            ws.cell(row=max_row + 1, column=2, value=current_date)
            ws.cell(row=max_row + 1, column=3, value=current_time)  # Add current time

            # Save the workbook
            wb.save(excel_file)

            print(f"Attendance marked for {object_name} on {current_date} at {current_time}.")

    # Load images and labels for all users
    all_images = []
    all_labels = []

    # Iterate over each folder in the data path
    for label, folder_name in enumerate(listdir(data_path)):
        folder_path = join(data_path, folder_name)
        # Check if it's a directory
        if not isfile(folder_path):
            # Load images and labels for the current user
            image_files = [f for f in listdir(folder_path) if isfile(join(folder_path, f))]
            images = [cv2.imread(join(folder_path, file), cv2.IMREAD_GRAYSCALE) for file in image_files]
            # Resize all images to the same dimensions
            resized_images = [cv2.resize(img, (200, 200)) for img in images]
            all_images.extend(resized_images)
            all_labels.extend([label] * len(resized_images))

    # Convert lists to numpy arrays
    Training_Data = np.array(all_images)
    Labels = np.array(all_labels)

    # Train the model
    model = cv2.face.LBPHFaceRecognizer_create()
    model.train(Training_Data, Labels)

    print("Dataset Model Training Complete!!!!!")

    # Load the face cascade classifier
    face_classifier = cv2.CascadeClassifier('C:/Users/E-TIME/Desktop/facerecog/haarcascade_frontalface_default.xml')

    def face_detector(img, size=0.5):
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        faces = face_classifier.detectMultiScale(gray, 1.3, 5)

        if len(faces) == 0:
            return img, []

        for (x, y, w, h) in faces:
            cv2.rectangle(img, (x, y), (x + w, y + h), (0, 255, 0), 2)
            roi = img[y:y + h, x:x + w]
            roi = cv2.resize(roi, (200, 200))

        return img, roi

    cap = cv2.VideoCapture(0)
    while True:
        ret, frame = cap.read()

        image, face = face_detector(frame)

        try:
            face = cv2.cvtColor(face, cv2.COLOR_BGR2GRAY)
            result = model.predict(face)

            if result[1] < 500:
                confidence = int(100 * (1 - (result[1]) / 300))
                if confidence > 80:
                    # Dynamically fetch the user name based on the label predicted by the model
                    object_name = listdir(data_path)[result[0]]  # Extracting user name from folder name
                    cv2.putText(image, object_name, (250, 450), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
                    mark_attendance(object_name)

                else:
                    cv2.putText(image, "Unknown", (250, 450), cv2.FONT_HERSHEY_COMPLEX, 1, (0, 0, 255), 2)

        except Exception as e:
            print(f"Error: {e}")
            cv2.putText(image, "Face Not Found", (250, 450), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 0, 0), 2)

        cv2.imshow('Face Cropper', image)

        if cv2.waitKey(1) == 13:
            break

    cap.release()
    cv2.destroyAllWindows()
    return render(request, 'face.html')
